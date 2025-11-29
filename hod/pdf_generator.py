import io
import os
from pathlib import Path
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.conf import settings

# Add try-except around the imports to handle missing models

try:
    from .models import CourseAllocation, FacultyAssignment
except ImportError:
    CourseAllocation = None
    FacultyAssignment = None


def _logo_path():
    return os.path.join(
        settings.BASE_DIR,
        "users",
        "static",
        "images",
        "malnad_college_of_engineering_logo.jpeg",
    )


def generate_hod_course_allocation_pdf(hod_assignment):
    """
    Generate PDF showing all courses allocated by HOD with faculty assignments.
    This is what HOD must complete (up to page 14).
    """
    buffer = io.BytesIO()
    
    # Use landscape for wide course mapping table
    p = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    
    # Header setup
    left_margin = 30 * mm
    right_margin = width - 30 * mm
    top_margin = height - 30 * mm
    
    y = top_margin
    
    # Logo & Header
    logo_path = _logo_path()
    if os.path.exists(logo_path):
        try:
            p.drawImage(logo_path, left_margin, y - 20 * mm, width=18 * mm, height=18 * mm, mask="auto")
        except Exception:
            pass
    
    text_x = left_margin + 22 * mm
    p.setFont("Helvetica-Bold", 14)
    p.drawString(text_x, y - 6 * mm, "MALNAD COLLEGE OF ENGINEERING, HASSAN")
    p.setFont("Helvetica", 9)
    p.drawString(text_x, y - 11 * mm, "(An Autonomous Institution — Hassan, Karnataka)")
    p.setFont("Helvetica-Bold", 11)
    p.drawCentredString(width / 2, y - 18 * mm, f"DEPARTMENT OF {hod_assignment.branch.code.upper()}")
    p.drawCentredString(width / 2, y - 22 * mm, "COURSE ALLOCATION & FACULTY ASSIGNMENT")
    
    p.setLineWidth(0.8)
    p.line(left_margin, y - 25 * mm, right_margin, y - 25 * mm)
    
    y = y - 30 * mm
    
    # Course details header
    p.setFont("Helvetica-Bold", 10)
    p.drawString(left_margin, y, f"Branch: {hod_assignment.branch.code} — {hod_assignment.branch.name}")
    y -= 8 * mm
    p.drawString(left_margin, y, f"Academic Year: 2024-2025")
    
    y -= 15 * mm
    
    # Get all courses and their faculty assignments
    courses = CourseAllocation.objects.filter(
        hod_assignment=hod_assignment
    ).order_by("course_code")
    
    if not courses.exists():
        p.setFont("Helvetica", 10)
        p.drawString(left_margin, y, "No courses allocated yet.")
        p.showPage()
        p.save()
        buffer.seek(0)
        return buffer
    
    # Build course table
    table_data = [
        ["#", "Course Code", "Course Title", "L-T-P", "Credits", "CIE/SEE", "Faculty - Role"]
    ]
    
    for idx, course in enumerate(courses, 1):
        assignments = FacultyAssignment.objects.filter(course_allocation=course)
        
        if assignments.exists():
            faculty_str = ", ".join([
                f"{fa.faculty.user.get_full_name()}\n({fa.role})"
                for fa in assignments
            ])
        else:
            faculty_str = "Not Assigned"
        
        table_data.append([
            str(idx),
            course.course_code,
            course.course_title[:25],  # Truncate if too long
            f"{course.teaching_hours_L}-{course.teaching_hours_T}-{course.teaching_hours_P}",
            str(course.credits),
            f"{course.cie_marks}/{course.see_marks}",
            faculty_str,
        ])
    
    # Create table
    table = Table(
        table_data,
        colWidths=[10 * mm, 20 * mm, 60 * mm, 18 * mm, 15 * mm, 18 * mm, 80 * mm]
    )
    
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("WORDWRAP", (0, 0), (-1, -1), True),
    ]))
    
    # Draw table
    table.wrapOn(p, width, height)
    table.drawOn(p, left_margin, y - 180 * mm)
    
    # Footer
    p.setFont("Helvetica-Oblique", 8)
    p.drawCentredString(width / 2, 15 * mm, "HOD Work Completion Report — Page 1 of 1")
    p.drawCentredString(width / 2, 10 * mm, "Generated via MCE Syllabus Maker © Malnad College of Engineering")
    
    p.showPage()
    p.save()
    buffer.seek(0)
    return buffer


def generate_hod_course_details_excel(hod_assignment):
    """
    Generate Excel sheet with all HOD's courses, schemes, and faculty assignments.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None
    
    from .models import CourseScheme
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses"
    
    # Header
    headers = ["#", "Course Code", "Title", "Category", "L", "T", "P", "Credits", 
               "CIE", "SEE", "Scheme Status", "Faculty Assigned"]
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    courses = CourseAllocation.objects.filter(hod_assignment=hod_assignment).order_by("course_code")
    
    for idx, course in enumerate(courses, 1):
        scheme = CourseScheme.objects.filter(course_allocation=course).first()
        assignments = FacultyAssignment.objects.filter(course_allocation=course)
        
        faculty_str = ", ".join([f"{fa.faculty.faculty_id}" for fa in assignments]) or "Not Assigned"
        
        ws.append([
            idx,
            course.course_code,
            course.course_title,
            course.course_category,
            course.teaching_hours_L,
            course.teaching_hours_T,
            course.teaching_hours_P,
            float(course.credits),
            course.cie_marks,
            course.see_marks,
            "✓" if scheme else "✗",
            faculty_str,
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 5
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 8
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 20
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_start_pages_pdf(branch, admission_year):
    """
    Generate a 7-page starting PDF for given branch and admission_year.
    File saved to MEDIA_ROOT/starting_pages/branch_<pk>_<year>.pdf
    Returns absolute filesystem path of generated PDF.
    Raises ImportError if reportlab not installed.
    """
    if canvas is None:
        raise ImportError("reportlab is required to generate PDFs. Install with: pip install reportlab")

    media_root = Path(getattr(settings, 'MEDIA_ROOT', settings.BASE_DIR))  # fallback to BASE_DIR
    out_dir = media_root / 'starting_pages'
    out_dir.mkdir(parents=True, exist_ok=True)

    filename = out_dir / f"starting_pages_branch_{branch.pk}_{admission_year}.pdf"
    c = canvas.Canvas(str(filename), pagesize=A4)
    width, height = A4

    # branch logo image path resolution - try model attribute names, else fall back to static image
    logo_path = None
    try:
        # common field names: logo, image, department_logo
        for fname in ('logo', 'image', 'department_logo'):
            if hasattr(branch, fname):
                val = getattr(branch, fname)
                if val:
                    # val may be a FileField/ImageField
                    try:
                        logo_path = val.path
                    except Exception:
                        # maybe .url or string - ignore
                        logo_path = None
                    break
    except Exception:
        logo_path = None

    # if no file path found, keep logo_path None; views/templates can use static if needed

    # Simple page template: branch name, admission year, page number
    left_margin = 20 * mm
    top_margin = height - 20 * mm

    for p in range(1, 8):
        # Draw header with branch name and year
        c.setFont("Helvetica-Bold", 18)
        branch_name = getattr(branch, 'name', str(branch))
        c.drawString(left_margin, top_margin, f"{branch_name} — Malnad College of Engineering")
        c.setFont("Helvetica", 12)
        c.drawString(left_margin, top_margin - 18, f"Admission Year: {admission_year}")
        c.drawString(left_margin, top_margin - 34, f"Generated Page: {p} / 7")

        # Draw a placeholder block (you can replace with richer layout / HTML->PDF later)
        c.setFont("Helvetica", 10)
        text_y = top_margin - 70
        lines = [
            "This is an automatically generated starting page.",
            "Replace this template with institution-specific content if needed.",
            "",
            "Contents (example):",
            "1. Cover & institution details",
            "2. College logo & approvals",
            "3. Scheme information",
            "4. Instructions for faculty",
            "",
            "Note: Branch logo, branch name and admission year are injected per branch."
        ]
        for ln in lines:
            c.drawString(left_margin, text_y, ln)
            text_y -= 14

        # draw branch logo on top-right if available
        if logo_path:
            try:
                # keep logo size moderate
                img_w = 40 * mm
                img_h = 40 * mm
                c.drawImage(str(logo_path), width - left_margin - img_w, top_margin - img_h/2, img_w, img_h, preserveAspectRatio=True, mask='auto')
            except Exception:
                # ignore image errors
                pass

        # page footer
        c.setFont("Helvetica-Oblique", 8)
        c.drawCentredString(width/2.0, 12 * mm, f"© Malnad College of Engineering — Starting Pages — Branch: {branch_name} — Year: {admission_year}")

        c.showPage()

    c.save()
    return str(filename)

def generate_scheme_pdf(courses, branch, year, semester, filename='scheme.pdf'):
    # Prepare file path
    pdf_path = os.path.join('media', filename)
    doc = SimpleDocTemplate(pdf_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Header
    elements.append(Paragraph('<b>MALNAD COLLEGE OF ENGINEERING, HASSAN</b>', styles['Title']))
    elements.append(Paragraph('DEPARTMENT OF %s' % branch.upper(), styles['Heading2']))
    elements.append(Paragraph('<b>THIRD SEMESTER</b>', styles['Heading3']))
    elements.append(Spacer(1, 12))

    # Table Data
    data = [
        ['Sl. No', 'Course Category', 'Course Code', 'Course Title',
         'L', 'T', 'P', 'Total', 'CIE', 'SEE', 'Total', 'Credits']
    ]
    for i, c in enumerate(courses, 1):
        data.append([
            i,
            c.category,
            c.course_code,
            c.course_title,
            getattr(c, 'l', ''),
            getattr(c, 't', ''),
            getattr(c, 'p', ''),
            getattr(c, 'total_hours', ''),
            getattr(c, 'cie', ''),
            getattr(c, 'see', ''),
            getattr(c, 'total_marks', ''),
            getattr(c, 'credits', ''),
        ])

    # Table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f3e6fa')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#4b0082')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    # Note
    elements.append(Paragraph(
        '<font color="#800080">Note: AEC, SEC, ETC courses are to be chosen suitably by the BOS of the programme</font>',
        styles['Normal']
    ))

    # Example for additional tables (like Engineering Science Course)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph('<b>Engineering Science Course (ESC/ETC/PLC)</b>', styles['Normal']))
    elements.append(Table([
        ['23IS306A', 'OOP with Java', '23IS306C', 'Discrete Mathematical Structures'],
        ['23IS306B', 'OOP with C++', '23IS306D', 'Graph Theory and Combinatorics'],
    ], colWidths=[60, 120, 60, 120]))

    doc.build(elements)
    return pdf_path